"""
Full Pipeline for ADL Refactoring Process

This script implements a clean two-agent pipeline:
1. ArchitectureDesigner - handles all refactoring tasks (initial, corrections, syntax fixes, misconfiguration handling)
2. EvaluatorAgent - handles all evaluations 

The pipeline uses a simple loop structure:
- Initial: architecture_designer.process() -> evaluator_agent.evaluate_adl()  
- Repair loop (up to 5 iterations): architecture_designer.process() -> evaluator_agent.evaluate_adl()

Now records separate usage for each LLM call (refactor, correct, fix_syntax, misconfiguration_handling)
"""

from agents.architecture_designer import ArchitectureDesigner
from agents.evaluator_agent import EvaluatorAgent

from openpyxl import load_workbook
import os
import pandas as pd
import time

def refactoring_process(adl_name, new_requirement):
    """
    Clean pipeline orchestrating refactoring and evaluation agents.
    
    Args:
        adl_name (str): Name of the ADL file
        new_requirement (str): New requirement to implement
        
    Returns:
        tuple: (detailed_result, agent_usage)
    """
    # Example: disable misconfiguration handling for testing
    # architecture_designer.disable_function("misconfiguration_handling")
    
    architecture_designer = ArchitectureDesigner()
    evaluator_agent = EvaluatorAgent()
    
    # Use Sonnet 4.6 with default thinking behavior (do not override thinking settings).
    architecture_designer.set_model("claude-sonnet-4-6")
    agent_usage = []
    
    # Initial refactoring
    current_adl, task_name = architecture_designer.process(adl_name=adl_name, new_requirement=new_requirement)
    separate_records = architecture_designer.get_separate_usage_records()
    agent_usage.extend(separate_records)
    
    # Initial evaluation
    detailed_result, non_existing_properties, result_enum, asserts = evaluator_agent.evaluate_adl(adl_name, current_adl, new_requirement)
    usage_record = evaluator_agent.record_latest_usage()
    if usage_record:
        usage_record['agent_name'] = usage_record['agent_name'] + " (initial_evaluation)"
        agent_usage.append(usage_record)
    
    # Repair loop (up to 5 iterations)
    for iteration in range(1, 6):
        # Check if we should continue processing
        if not architecture_designer.skill_selector.should_continue(
            architecture_designer.skill_selector.determine_task(result_enum, non_existing_properties)
        ):
            break
            
        # Repair step
        current_adl, task_name = architecture_designer.process(
            current_adl=current_adl,
            result_enum=result_enum,
            non_existing_properties=non_existing_properties
        )
        separate_records = architecture_designer.get_separate_usage_records()
        agent_usage.extend(separate_records)
        
        # Re-evaluate
        detailed_result, non_existing_properties, result_enum, asserts = evaluator_agent.evaluate_adl(adl_name, current_adl, new_requirement)
        usage_record = evaluator_agent.record_latest_usage()
        if usage_record:
            usage_record['agent_name'] = usage_record['agent_name'] + f" (iteration_{iteration}_evaluation)"
            agent_usage.append(usage_record)
    
    print("agent_usage:", agent_usage)
    
    return detailed_result, agent_usage, asserts

def append_to_excel(filename, df):
    if os.path.exists(filename):
        with pd.ExcelWriter(filename, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            book = load_workbook(filename)
            sheet = writer.sheets['Sheet1']
            start_row = sheet.max_row
            df.to_excel(writer, index=False, header=False, startrow=start_row)
    else:
        df.to_excel(filename, index=False, engine="openpyxl")


def filter_requirements_for_system(df: pd.DataFrame, system_name: str) -> pd.DataFrame:
    """
    Return only the requirement rows matching the given System name.
    This keeps main() simple while letting us focus runs on a single ADL.
    """
    if "System" not in df.columns:
        return df
    mask = df["System"].astype(str).str.strip().str.lower() == system_name.lower()
    return df[mask].reset_index(drop=True)

def main():
    input_file = "new_clustered_requirements.xlsx"
    output_file = "Results/Fullpipeline/claude-4.6-sonnet.xlsx"
    start_index = 0
    end_index = 120
    
    df = pd.read_excel(input_file, engine="openpyxl")
    df.columns = df.columns.astype(str).str.strip()
    df = df.iloc[start_index:end_index]


    # Prepare Excel file with headers if not exist
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    if not os.path.exists(output_file):
        df_empty = pd.DataFrame(columns=["ADL File Name", "New Requirement", "Run ID", "Execution Result", "Runtime (s)","Assert Statements", "Agent Usage"])
        df_empty.to_excel(output_file, index=False, engine="openpyxl")

    for index, row in df.iterrows():
        adl_name = row['System']
        new_requirement = row['Requirement']

        print(f"ADL: {adl_name}")
        print(f"New Requirement: {new_requirement}")
        print("-" * 40)
        
        start_time = time.time()
        detailed_result, agent_usage, asserts = refactoring_process(adl_name, new_requirement)
        end_time = time.time()
        runtime_seconds = end_time - start_time

        result_df = pd.DataFrame([[
            adl_name,
            new_requirement,
            1,  # or another run ID
            detailed_result,
            runtime_seconds,
            asserts,
            agent_usage
        ]], columns=["ADL File Name", "New Requirement", "Run ID", "Execution Result", "Runtime (s)", "Assert Statements", "Agent Usage"])
        append_to_excel(output_file, result_df)

if __name__ == "__main__":
    main()